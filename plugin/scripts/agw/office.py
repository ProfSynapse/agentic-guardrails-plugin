"""Controlled in-place edits to Office files (docx/xlsx/pptx).

The sanctioned alternative to ad-hoc interpreter one-liners: every mutating
operation archives a pre-image snapshot before touching the file, so the edit
is reversible with `agw restore` no matter what the library does to the file.

The built-in OOXML backend covers Word/PowerPoint text operations and surgical
Excel cell edits without third-party packages. Advanced workbook operations use
optional openpyxl and report exactly what to install when it is missing.
"""
from __future__ import annotations

import os

from core import store
import office_ooxml
import office_tx


class OfficeError(Exception):
    """Operation failed in a way the caller should report verbatim."""


class MissingLibrary(OfficeError):
    def __init__(self, lib: str, pip_name: str, ext: str):
        super().__init__(
            f"{ext} support needs the optional '{pip_name}' package in the "
            "Python runtime selected by agw; Office libraries are not bundled"
        )
        self.lib, self.pip_name = lib, pip_name


def _openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        raise MissingLibrary("openpyxl", "openpyxl", ".xlsx")


def capabilities() -> dict:
    try:
        _openpyxl()
        xlsx_advanced = True
    except MissingLibrary:
        xlsx_advanced = False
    return {
        "docx": True,
        "pptx": True,
        "xlsx": True,
        "xlsx_advanced": xlsx_advanced,
    }


def _ext(path: str) -> str:
    return os.path.splitext(path)[1].lower()


def preservation_info(path: str) -> dict:
    """Return read-only OOXML preservation risk information."""
    office_tx._package_preflight(path, mutating=False)
    risks = office_tx.inspect_preservation_risks(path)
    return {"safe_to_mutate": not risks, "risks": risks}


def _snapshot(path: str, op: str) -> dict:
    return store.archive_file(path, mode="copy", dedupe=True,
                              reason=f"pre-image before agw office {op}")


def find_matches(path: str, find: str) -> list:
    """List every occurrence with a 1-based index, location, and context —
    what an agent uses to choose --nth or confirm --all."""
    if not find:
        raise OfficeError("--find must not be empty")
    try:
        if _ext(path) == ".docx":
            return office_ooxml.word_find_matches(path, find)
        if _ext(path) == ".pptx":
            return office_ooxml.presentation_find_matches(path, find)
    except office_ooxml.OoxmlError as exc:
        raise OfficeError(str(exc)) from exc
    raise OfficeError("replace-text supports .docx and .pptx")


# --- read operations (no snapshot needed) -------------------------------------

def get_text(path: str) -> str:
    preservation_info(path)
    ext = _ext(path)
    try:
        if ext == ".docx":
            return office_ooxml.word_get_text(path)
        if ext == ".pptx":
            return office_ooxml.presentation_get_text(path)
    except office_ooxml.OoxmlError as exc:
        raise OfficeError(str(exc)) from exc
    raise OfficeError(f"get-text supports .docx and .pptx (for {ext or 'this file'}, "
                      "use `agw checkout` / `agw convert`)")


def info(path: str) -> dict:
    ext = _ext(path)
    if ext == ".docx":
        preservation = preservation_info(path)
        try:
            return {**office_ooxml.word_info(path), "preservation": preservation}
        except office_ooxml.OoxmlError as exc:
            raise OfficeError(str(exc)) from exc
    if ext in (".xlsx", ".xlsm"):
        import office_excel
        data = office_excel.workbook_info(path)
        data["type"] = ext.lstrip(".")
        if ext == ".xlsm":
            manifest = office_tx.package_preservation_manifest(path)
            data["macro_preservation"] = {
                "schema": manifest["schema"],
                "protected_part_count": manifest["protected_part_count"],
                "categories": manifest["categories"],
                "file_sha256": manifest["file_sha256"],
            }
        # Preserve the legacy sheet-name mapping while returning richer data.
        data["sheet_list"] = data.pop("sheets")
        data["sheets"] = {
            item["name"]: {
                key: value for key, value in item.items() if key != "name"
            }
            for item in data["sheet_list"]
        }
        return data
    if ext == ".pptx":
        preservation = preservation_info(path)
        try:
            return {**office_ooxml.presentation_info(path),
                    "preservation": preservation}
        except office_ooxml.OoxmlError as exc:
            raise OfficeError(str(exc)) from exc
    raise OfficeError(f"unsupported extension: {ext or 'none'}")


# --- write operations (snapshot first, always) ---------------------------------

def replace_text(path: str, find: str, replace: str,
                 all_matches: bool = False, nth: int = 0) -> dict:
    """Replace occurrences of `find`. Same contract as a code editor's
    find/replace tool: a non-unique match without explicit targeting is an
    error, not a mass edit.

    - unique match           -> replaced
    - multiple matches       -> OfficeError unless all_matches or nth is given
    - nth=N (1-based, doc order) -> replace only that occurrence
    """
    if not find:
        raise OfficeError("--find must not be empty")
    if all_matches and nth:
        raise OfficeError("--all and --nth are mutually exclusive")
    total = len(find_matches(path, find))
    if total == 0:
        return {"replacements": 0, "matches": 0}
    if nth:
        if not 1 <= nth <= total:
            raise OfficeError(f"--nth {nth} out of range: {total} match(es)")
        targets = {nth}
    elif total > 1 and not all_matches:
        preview = "; ".join(f"#{m['n']} ({m['where']}) ...{m['context']}..."
                            for m in find_matches(path, find)[:5])
        raise OfficeError(
            f"{total} matches for {find!r} — refusing an ambiguous replace. "
            f"Use --all for every occurrence, --nth N for one, or a longer "
            f"--find that is unique. First matches: {preview}")
    else:
        targets = None  # all
    result_holder = {}

    def plan(_live):
        return office_tx.MutationPlan(
            "replace-text",
            {"replacements": len(targets) if targets is not None else total,
             "matches": total},
            {"affected": len(targets) if targets is not None else total},
        )

    def apply(stage, _plan):
        try:
            if _ext(stage) == ".docx":
                result_holder["replacements"] = office_ooxml.word_replace(
                    stage, find, replace, targets
                )
            else:
                result_holder["replacements"] = office_ooxml.presentation_replace(
                    stage, find, replace, targets
                )
        except office_ooxml.OoxmlError as exc:
            raise OfficeError(str(exc)) from exc

    def validate(stage, _plan):
        remaining = len(find_matches(stage, find))
        expected = total - result_holder["replacements"]
        if remaining != expected:
            raise OfficeError("staged replace-text validation failed")
        return {"replacements": result_holder["replacements"], "matches": total}

    try:
        result = office_tx.execute_mutation(
            path, operation="replace-text", plan=plan,
            apply=apply, validate=validate,
        )
    except office_tx.TransactionError as exc:
        raise OfficeError(str(exc)) from exc
    result["snapshot_version"] = result.pop("snapshot", None)
    return result


def _coerce(value: str, force_text: bool):
    if force_text:
        return value
    if value.startswith("="):
        return value  # formula
    for caster in (int, float):
        try:
            return caster(value)
        except ValueError:
            continue
    if value.lower() in ("true", "false"):
        return value.lower() == "true"
    return value


def set_cell(path: str, sheet: str, cell: str, value: str,
             force_text: bool = False, expected_sha256: str = "",
             dry_run: bool = False) -> dict:
    office_tx._package_preflight(path, mutating=False)
    risks = office_tx.inspect_preservation_risks(path)
    macro_enabled = _ext(path) == ".xlsm"
    if _ext(path) in (".xlsx", ".xlsm"):
        import office_surgical
        state = {"new": _coerce(value, force_text)}

        def surgical_plan(live):
            inspected = office_surgical.inspect_cell(live, sheet, cell)
            state.update(inspected)
            old = inspected["value"]
            return office_tx.MutationPlan(
                "set-cell",
                {"sheet": sheet, "cell": inspected["coordinate"],
                 "old": old, "new": state["new"],
                 "adapter": "ooxml-surgical"},
                {"affected": int(old != state["new"])},
                changed=old != state["new"],
            )

        def surgical_apply(stage, _plan):
            office_surgical.set_cell(stage, sheet, state["coordinate"], state["new"])

        def surgical_validate(stage, _plan):
            return office_surgical.verify_cell(
                stage, sheet, state["coordinate"], state["new"]
            )

        def surgical_preservation(before, after, _operation):
            return office_surgical.verify_preservation(
                before, after, state["part"]
            )

        try:
            result = office_tx.execute_mutation(
                path, operation="set-cell", plan=surgical_plan,
                apply=surgical_apply, validate=surgical_validate,
                expected_sha256=expected_sha256 or None, dry_run=dry_run,
                allow_preservation_risks=True,
                allow_macro_enabled=macro_enabled,
                preservation_validator=surgical_preservation,
            )
        except (office_tx.TransactionError,
                office_surgical.SurgicalCellError) as exc:
            raise OfficeError(str(exc)) from exc
        result["snapshot_version"] = result.pop("snapshot", None)
        result["preserved_risks"] = risks
        if macro_enabled:
            result["macro_preservation"] = (
                office_tx.package_preservation_manifest(path)
            )["categories"]
        return result

    openpyxl = _openpyxl()
    state = {}

    def plan(live):
        wb = openpyxl.load_workbook(live)
        try:
            if sheet not in wb.sheetnames:
                raise OfficeError(
                    f"no sheet named {sheet!r} (have: {', '.join(wb.sheetnames)})"
                )
            old = wb[sheet][cell].value
            new = _coerce(value, force_text)
            state.update({"old": old, "new": new})
            return office_tx.MutationPlan(
                "set-cell", {"sheet": sheet, "cell": cell, "old": old, "new": new},
                {"affected": int(old != new)}, changed=old != new,
            )
        finally:
            wb.close()

    def apply(stage, _plan):
        wb = openpyxl.load_workbook(stage)
        wb[sheet][cell] = state["new"]
        wb.save(stage)
        wb.close()

    def validate(stage, _plan):
        wb = openpyxl.load_workbook(stage, data_only=False)
        try:
            if wb[sheet][cell].value != state["new"]:
                raise OfficeError("staged set-cell validation failed")
        finally:
            wb.close()
        return {}

    try:
        result = office_tx.execute_mutation(
            path, operation="set-cell", plan=plan, apply=apply, validate=validate,
            expected_sha256=expected_sha256 or None, dry_run=dry_run,
        )
    except office_tx.TransactionError as exc:
        raise OfficeError(str(exc)) from exc
    result["snapshot_version"] = result.pop("snapshot", None)
    return result


def append_rows(path: str, sheet: str, rows: list, force_text: bool = False) -> dict:
    openpyxl = _openpyxl()
    state = {}

    def plan(live):
        wb = openpyxl.load_workbook(live)
        try:
            if sheet not in wb.sheetnames:
                raise OfficeError(
                    f"no sheet named {sheet!r} (have: {', '.join(wb.sheetnames)})"
                )
            state["rows_before"] = wb[sheet].max_row
            return office_tx.MutationPlan(
                "append-rows", {"sheet": sheet, "appended": len(rows)},
                {"affected": len(rows)}, changed=bool(rows),
            )
        finally:
            wb.close()

    def apply(stage, _plan):
        wb = openpyxl.load_workbook(stage)
        ws = wb[sheet]
        for row in rows:
            ws.append([_coerce(str(v), force_text) if v is not None else None
                       for v in row])
        wb.save(stage)
        wb.close()

    def validate(stage, _plan):
        wb = openpyxl.load_workbook(stage)
        try:
            rows_now = wb[sheet].max_row
            if rows_now != state["rows_before"] + len(rows):
                raise OfficeError("staged append-rows validation failed")
            return {"rows_now": rows_now}
        finally:
            wb.close()

    try:
        result = office_tx.execute_mutation(
            path, operation="append-rows", plan=plan,
            apply=apply, validate=validate,
        )
    except office_tx.TransactionError as exc:
        raise OfficeError(str(exc)) from exc
    result["snapshot_version"] = result.pop("snapshot", None)
    return result
