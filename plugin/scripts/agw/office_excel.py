"""Structured, schema-agnostic Excel reads and named-table mutations."""
from __future__ import annotations

import copy
import datetime as dt
import hashlib
import json
import re
import warnings
from typing import Optional

from core import store
try:
    from openpyxl.formula.translate import Translator
    from openpyxl.utils import get_column_letter, range_boundaries
    from openpyxl.worksheet.table import Table, TableFormula, TableStyleInfo
except ImportError:  # optional dependency; _load() reports the actionable error
    Translator = get_column_letter = range_boundaries = None
    Table = TableFormula = TableStyleInfo = None

import office_tx

MAX_INSPECTED_CELLS = 2_000_000
DEFAULT_LIMIT = 50
MAX_LIMIT = 500
MAX_RETURNED_CELLS = 10_000
_EXTERNAL_FORMULA_RE = re.compile(
    r"(?i)(?:\[[^\]]+\.(?:xlsx|xlsm|xlsb|xls)\]|\[\d+\][^!]*!)"
)


class ExcelError(Exception):
    pass


class ExcelConflict(ExcelError):
    error_code = "uniqueness_conflict"

    def __init__(self, message: str, details: dict):
        self.details = details
        super().__init__(f"CONFLICT: {message}")


def _openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError as exc:
        raise ExcelError(".xlsx support needs the 'openpyxl' package") from exc


def _load(path: str, *, data_only: bool = False):
    kwargs = {
        "data_only": data_only,
        "read_only": False,
        "keep_links": True,
    }
    try:
        return _openpyxl().load_workbook(path, rich_text=True, **kwargs)
    except TypeError:
        return _openpyxl().load_workbook(path, **kwargs)


def _load_for_read(path: str, *, data_only: bool = False):
    with warnings.catch_warnings(record=True) as captured:
        warnings.simplefilter("always")
        workbook = _load(path, data_only=data_only)
    risks = office_tx.inspect_preservation_risks(path)
    risks.extend(office_tx._warning_risks(captured, "read"))
    return workbook, risks


def _preservation_result(risks: list[dict]) -> dict:
    return {"safe_to_mutate": not risks, "risks": risks}


def _actual_bounds(ws) -> dict:
    cells = getattr(ws, "_cells", None)
    if not isinstance(cells, dict):
        raise ExcelError("installed openpyxl lacks the supported sparse-cell interface")
    if len(cells) > MAX_INSPECTED_CELLS:
        raise ExcelError("worksheet has too many instantiated cells to inspect safely")
    min_row = min_col = max_row = max_col = None
    count = 0
    for cell in cells.values():
        if cell.value is None:
            continue
        count += 1
        min_row = cell.row if min_row is None else min(min_row, cell.row)
        max_row = cell.row if max_row is None else max(max_row, cell.row)
        min_col = cell.column if min_col is None else min(min_col, cell.column)
        max_col = cell.column if max_col is None else max(max_col, cell.column)
    used = None
    rows = cols = 0
    if count:
        used = (
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(max_col)}{max_row}"
        )
        rows = max_row - min_row + 1
        cols = max_col - min_col + 1
    return {
        "used_range": used,
        "rows": rows,
        "cols": cols,
        "nonempty_cells": count,
        "worksheet_dimension": ws.calculate_dimension(),
    }


def _table_headers(ws, table):
    min_col, min_row, max_col, _ = range_boundaries(table.ref)
    headers = [ws.cell(min_row, col).value for col in range(min_col, max_col + 1)]
    if any(value is None or str(value).strip() == "" for value in headers):
        raise ExcelError(f"table {table.displayName!r} has blank headers")
    names = [str(value) for value in headers]
    folded = [value.casefold() for value in names]
    if len(set(folded)) != len(folded):
        raise ExcelError(f"table {table.displayName!r} has duplicate headers")
    return names


def _find_table(wb, name: str, sheet: str = ""):
    matches = []
    for ws in wb.worksheets:
        if sheet and ws.title.casefold() != sheet.casefold():
            continue
        for table in ws.tables.values():
            display = getattr(table, "displayName", None) or getattr(table, "name", "")
            if str(display).casefold() == name.casefold():
                matches.append((ws, table))
    if not matches:
        raise ExcelError(f"no table named {name!r}")
    if len(matches) != 1:
        where = ", ".join(ws.title for ws, _ in matches)
        raise ExcelError(f"table name {name!r} is ambiguous across sheets: {where}")
    return matches[0]


def _totals_count(table) -> int:
    count = getattr(table, "totalsRowCount", None)
    if count is not None:
        return int(count)
    return 1 if getattr(table, "totalsRowShown", False) else 0


def workbook_info(path: str, scope: str = "") -> dict:
    office_tx._package_preflight(path, mutating=False)
    wb, risks = _load_for_read(path)
    try:
        sheets = []
        for ws in wb.worksheets:
            bounds = _actual_bounds(ws)
            tables = []
            for table in ws.tables.values():
                headers = _table_headers(ws, table)
                min_col, min_row, max_col, max_row = range_boundaries(table.ref)
                totals = _totals_count(table)
                tables.append({
                    "name": table.displayName,
                    "ref": table.ref,
                    "headers": headers,
                    "rows": max(0, max_row - min_row - totals),
                    "totals_row": bool(totals),
                })
            item = {"name": ws.title, **bounds}
            if scope in ("", "tables"):
                item["tables"] = tables
            if scope == "":
                item["merged_ranges"] = [str(value) for value in ws.merged_cells.ranges]
            sheets.append(item)
        names = []
        if scope in ("", "names"):
            try:
                values = wb.defined_names.values()
            except AttributeError:
                values = wb.defined_names.definedName
            for value in values:
                names.append({
                    "name": value.name,
                    "attr_text": value.attr_text,
                    "local_sheet_id": value.localSheetId,
                })
        result = {
            "type": "xlsx", "hash": store.file_sha256(path),
            "preservation": _preservation_result(risks),
        }
        if scope != "names":
            result["sheets"] = sheets
        if scope in ("", "names"):
            result["defined_names"] = names
        return result
    finally:
        wb.close()


def _json_value(value, cell=None):
    if isinstance(value, dt.datetime) and cell is not None:
        fmt = (cell.number_format or "").lower()
        if not any(token in fmt for token in ("h", "s")):
            return value.date().isoformat()
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    return value


def read_table(
    path: str,
    table_name: str,
    *,
    sheet: str = "",
    columns: Optional[list[str]] = None,
    where: Optional[dict] = None,
    offset: int = 0,
    limit: int = DEFAULT_LIMIT,
    values_only: bool = False,
    include_formulas: bool = False,
) -> dict:
    if offset < 0 or limit < 1 or limit > MAX_LIMIT:
        raise ExcelError(f"offset must be >= 0 and limit must be 1..{MAX_LIMIT}")
    office_tx._package_preflight(path, mutating=False)
    wb, risks = _load_for_read(
        path, data_only=(values_only and not include_formulas)
    )
    cached_wb = None
    try:
        ws, table = _find_table(wb, table_name, sheet)
        cached_ws = None
        if include_formulas:
            cached_wb, _cached_risks = _load_for_read(path, data_only=True)
            cached_ws, _cached_table = _find_table(cached_wb, table_name, sheet)
        headers = _table_headers(ws, table)
        selected = columns or headers
        unknown = [name for name in selected if name not in headers]
        if unknown:
            raise ExcelError(f"unknown table column(s): {', '.join(unknown)}")
        if len(selected) != len(set(selected)):
            raise ExcelError("selected columns must be unique")
        where = where or {}
        unknown_where = [name for name in where if name not in headers]
        if unknown_where:
            raise ExcelError(f"unknown filter column(s): {', '.join(unknown_where)}")
        if len(selected) * limit > MAX_RETURNED_CELLS:
            raise ExcelError("requested table page exceeds the returned-cell limit")
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        data_end = max_row - _totals_count(table)
        indexes = {name: min_col + headers.index(name) for name in headers}
        matched = []
        value_ws = cached_ws or ws
        for row_no in range(min_row + 1, data_end + 1):
            if all(_json_value(value_ws.cell(row_no, indexes[key]).value,
                               value_ws.cell(row_no, indexes[key])) == expected
                   for key, expected in where.items()):
                matched.append(row_no)
        page = matched[offset:offset + limit]
        rows = []
        for row_no in page:
            row = []
            for name in selected:
                formula_cell = ws.cell(row_no, indexes[name])
                if include_formulas:
                    cached_cell = cached_ws.cell(row_no, indexes[name])
                    formula = formula_cell.value if (
                        formula_cell.data_type == "f"
                        or (isinstance(formula_cell.value, str)
                            and formula_cell.value.startswith("="))
                    ) else None
                    row.append({
                        "value": _json_value(
                            cached_cell.value if formula else formula_cell.value,
                            cached_cell if formula else formula_cell,
                        ),
                        "formula": formula,
                    })
                else:
                    row.append(_json_value(formula_cell.value, formula_cell))
            rows.append(row)
        return {
            "hash": store.file_sha256(path),
            "table": table.displayName,
            "sheet": ws.title,
            "ref": table.ref,
            "headers": selected,
            "rows": rows,
            "offset": offset,
            "returned": len(rows),
            "more": offset + len(rows) < len(matched),
            "row_count": max(0, data_end - min_row),
            "preservation": _preservation_result(risks),
            **({"cached_values_may_be_stale": True}
               if values_only or include_formulas else {}),
            **({"formula_mode": "value_and_formula"} if include_formulas else {}),
        }
    finally:
        if cached_wb is not None:
            cached_wb.close()
        wb.close()


def read_range(
    path: str,
    sheet: str,
    cell_range: str,
    *,
    include_formulas: bool = False,
) -> dict:
    office_tx._package_preflight(path, mutating=False)
    try:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    except (TypeError, ValueError) as exc:
        raise ExcelError("range must be a finite A1 rectangle") from exc
    cells = (max_col - min_col + 1) * (max_row - min_row + 1)
    if cells < 1 or cells > MAX_RETURNED_CELLS:
        raise ExcelError(
            f"range must contain 1..{MAX_RETURNED_CELLS} cells"
        )
    wb, risks = _load_for_read(path, data_only=False)
    cached_wb = None
    try:
        ws = _sheet_by_name(wb, sheet)
        if ws is None:
            raise ExcelError(f"no worksheet named {sheet!r}")
        cached_ws = None
        if include_formulas:
            cached_wb, _cached_risks = _load_for_read(path, data_only=True)
            cached_ws = _sheet_by_name(cached_wb, sheet)
        rows = []
        for row_no in range(min_row, max_row + 1):
            row = []
            for col_no in range(min_col, max_col + 1):
                cell = ws.cell(row_no, col_no)
                if include_formulas:
                    cached = cached_ws.cell(row_no, col_no)
                    formula = cell.value if (
                        cell.data_type == "f"
                        or (isinstance(cell.value, str) and cell.value.startswith("="))
                    ) else None
                    row.append({
                        "value": _json_value(
                            cached.value if formula else cell.value,
                            cached if formula else cell,
                        ),
                        "formula": formula,
                    })
                else:
                    row.append(_json_value(cell.value, cell))
            rows.append(row)
        return {
            "hash": store.file_sha256(path), "sheet": ws.title,
            "range": cell_range, "rows": rows, "cell_count": cells,
            "preservation": _preservation_result(risks),
            **({"formula_mode": "value_and_formula",
                "cached_values_may_be_stale": True} if include_formulas else {}),
        }
    finally:
        if cached_wb is not None:
            cached_wb.close()
        wb.close()


def validate_formulas(path: str, *, offset: int = 0,
                      limit: int = DEFAULT_LIMIT) -> dict:
    if offset < 0 or limit < 1 or limit > MAX_LIMIT:
        raise ExcelError(f"offset must be >= 0 and limit must be 1..{MAX_LIMIT}")
    office_tx._package_preflight(path, mutating=False)
    wb, risks = _load_for_read(path, data_only=False)
    cached_wb = None
    try:
        cached_wb, _cached_risks = _load_for_read(path, data_only=True)
        formulas = []
        for ws in wb.worksheets:
            cached_ws = _sheet_by_name(cached_wb, ws.title)
            bounds = _actual_bounds(ws)
            if not bounds.get("used_range"):
                continue
            min_col, min_row, max_col, max_row = range_boundaries(
                bounds["used_range"]
            )
            for row_no in range(min_row, max_row + 1):
                for col_no in range(min_col, max_col + 1):
                    cell = ws.cell(row_no, col_no)
                    if cell.data_type != "f" and not (
                            isinstance(cell.value, str) and cell.value.startswith("=")):
                        continue
                    cached = cached_ws.cell(row_no, col_no)
                    formulas.append({
                        "sheet": ws.title, "cell": cell.coordinate,
                        "formula": cell.value,
                        "cached_value": _json_value(cached.value, cached),
                        "cached_value_missing": cached.value is None,
                        "external_reference": (
                            isinstance(cell.value, str)
                            and bool(_EXTERNAL_FORMULA_RE.search(cell.value))
                        ),
                    })
        page = formulas[offset:offset + limit]
        missing = sum(item["cached_value_missing"] for item in formulas)
        external = sum(item["external_reference"] for item in formulas)
        return {
            "hash": store.file_sha256(path),
            "validation": "structural_only",
            "calculation_performed": False,
            "formula_count": len(formulas),
            "missing_cached_values": missing,
            "external_references": external,
            "formulas": page,
            "offset": offset, "returned": len(page),
            "more": offset + len(page) < len(formulas),
            "valid": external == 0,
            "preservation": _preservation_result(risks),
        }
    finally:
        if cached_wb is not None:
            cached_wb.close()
        wb.close()


def _column_is_date(ws, row: int, col: int) -> bool:
    cell = ws.cell(row, col)
    if getattr(cell, "is_date", False):
        return True
    fmt = (cell.number_format or "").lower()
    return any(token in fmt for token in ("yy", "dd", "mm", "hh", "ss"))


def _coerce(value, *, date_column: bool = False, coerce_iso_dates: bool = False):
    if isinstance(value, dict):
        if set(value) != {"$formula"} or not isinstance(value["$formula"], str):
            raise ExcelError("typed values support only {'$formula': '=...'}")
        formula = value["$formula"]
        if not formula.startswith("="):
            raise ExcelError("explicit formula must start with '='")
        lowered = formula.lower()
        if (_EXTERNAL_FORMULA_RE.search(formula)
                or any(token in lowered for token in ("dde(", "http:", "https:"))):
            raise ExcelError("external-reference formulas are unsupported")
        return formula
    if isinstance(value, str) and value.startswith("="):
        raise ExcelError(
            "formula-like strings require the typed {'$formula': '=...'} form"
        )
    if isinstance(value, str) and (date_column or coerce_iso_dates):
        try:
            if "t" in value.lower() or " " in value:
                parsed = dt.datetime.fromisoformat(value)
                if parsed.tzinfo is not None:
                    raise ExcelError("timezone-aware Excel datetimes are unsupported")
                return parsed
            return dt.date.fromisoformat(value)
        except ValueError:
            pass
    return value


def _copy_cell_style(source, target):
    target._style = copy.copy(source._style)
    if source.has_style:
        target.number_format = source.number_format


def _table_column_formula(table, offset: int) -> str:
    columns = getattr(table, "tableColumns", None) or []
    if offset >= len(columns):
        return ""
    formula = getattr(columns[offset], "calculatedColumnFormula", None)
    text = getattr(formula, "attr_text", None) or getattr(formula, "text", None) or ""
    return "=" + text.lstrip("=") if text else ""


def _validated_headers(values) -> list[str]:
    if not isinstance(values, list) or not values:
        raise ExcelError("headers must be a non-empty JSON array")
    if any(not isinstance(value, str) or not value.strip() for value in values):
        raise ExcelError("table headers must be non-empty strings")
    headers = [value.strip() for value in values]
    folded = [value.casefold() for value in headers]
    if len(set(folded)) != len(folded):
        raise ExcelError("table headers must be unique")
    return headers


def _column_metadata(metadata, headers: list[str]) -> dict[str, dict]:
    if metadata in (None, {} , []):
        return {}
    normalized = {}
    if isinstance(metadata, dict):
        items = []
        for name, value in metadata.items():
            if not isinstance(value, dict):
                raise ExcelError("column metadata values must be JSON objects")
            items.append({"name": name, **value})
    elif isinstance(metadata, list):
        items = metadata
    else:
        raise ExcelError("column metadata must be an object or array")
    for item in items:
        if not isinstance(item, dict):
            raise ExcelError("each column metadata entry must be an object")
        unknown_fields = set(item) - {"name", "number_format", "formula"}
        if unknown_fields:
            raise ExcelError(
                "unknown column metadata field(s): " + ", ".join(sorted(unknown_fields))
            )
        name = item.get("name")
        if name not in headers:
            raise ExcelError(f"unknown metadata column: {name!r}")
        if name in normalized:
            raise ExcelError(f"duplicate column metadata: {name!r}")
        value = {key: item[key] for key in ("number_format", "formula") if key in item}
        if "number_format" in value and (
                not isinstance(value["number_format"], str)
                or not value["number_format"].strip()):
            raise ExcelError("number_format metadata must be a non-empty string")
        if "formula" in value:
            if not isinstance(value["formula"], str):
                raise ExcelError("formula metadata must be a string")
            _coerce({"$formula": value["formula"]})
        normalized[name] = value
    return normalized


def _sheet_by_name(wb, name: str):
    matches = [ws for ws in wb.worksheets if ws.title.casefold() == name.casefold()]
    if len(matches) > 1:
        raise ExcelError(f"worksheet name {name!r} is ambiguous")
    return matches[0] if matches else None


def _table_matches(wb, name: str):
    matches = []
    for ws in wb.worksheets:
        for table in ws.tables.values():
            display = getattr(table, "displayName", None) or getattr(table, "name", "")
            if str(display).casefold() == name.casefold():
                matches.append((ws, table))
    return matches


def _normalized_range(cell_range: str) -> tuple[str, tuple[int, int, int, int]]:
    try:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    except (TypeError, ValueError) as exc:
        raise ExcelError(f"invalid rectangular range: {cell_range!r}") from exc
    if not all(isinstance(value, int) and value > 0
               for value in (min_col, min_row, max_col, max_row)):
        raise ExcelError(f"invalid rectangular range: {cell_range!r}")
    normalized = (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max_col)}{max_row}"
    )
    return normalized, (min_col, min_row, max_col, max_row)


def _ranges_overlap(first, second) -> bool:
    a_min_col, a_min_row, a_max_col, a_max_row = first
    b_min_col, b_min_row, b_max_col, b_max_row = second
    return not (
        a_max_col < b_min_col or b_max_col < a_min_col
        or a_max_row < b_min_row or b_max_row < a_min_row
    )


def _range_digest(ws, bounds) -> str:
    min_col, min_row, max_col, max_row = bounds
    payload = [
        [
            (ws.cell(row, col).data_type, ws.cell(row, col).value)
            for col in range(min_col, max_col + 1)
        ]
        for row in range(min_row, max_row + 1)
    ]
    return hashlib.sha256(
        json.dumps(payload, ensure_ascii=False, default=str,
                   separators=(",", ":")).encode("utf-8")
    ).hexdigest()


def _apply_table_metadata(ws, table, headers: list[str], metadata: dict[str, dict],
                          bounds, *, refuse_conflicts: bool) -> None:
    min_col, min_row, max_col, max_row = bounds
    if not getattr(table, "tableColumns", None):
        table._initialise_columns()
    for offset, column in enumerate(table.tableColumns):
        if offset < len(headers):
            column.name = headers[offset]
    for name, item in metadata.items():
        offset = headers.index(name)
        col = min_col + offset
        if "number_format" in item:
            for row_no in range(min_row + 1, max_row + 1):
                ws.cell(row_no, col).number_format = item["number_format"]
        if "formula" in item:
            formula = item["formula"]
            origin = f"{get_column_letter(col)}{min_row + 1}"
            for row_no in range(min_row + 1, max_row + 1):
                cell = ws.cell(row_no, col)
                expected = Translator(formula, origin=origin).translate_formula(cell.coordinate)
                if cell.value not in (None, expected) and refuse_conflicts:
                    raise ExcelError(
                        f"formula metadata would overwrite populated cell {cell.coordinate}"
                    )
                cell.value = expected
            table.tableColumns[offset].calculatedColumnFormula = TableFormula(
                attr_text=formula[1:]
            )


def ensure_table(
    path: str,
    table_name: str,
    *,
    sheet: str,
    headers: Optional[list[str]] = None,
    cell_range: str = "",
    style: str = "",
    columns=None,
    create_sheet: bool = False,
    expected_sha256: str = "",
    dry_run: bool = False,
) -> dict:
    """Create or verify a schema-agnostic named Excel table."""
    if not table_name or not isinstance(table_name, str):
        raise ExcelError("--table is required")
    if not sheet or not isinstance(sheet, str):
        raise ExcelError("--sheet is required")
    requested_headers = _validated_headers(headers) if headers is not None else None
    if not cell_range and requested_headers is None:
        raise ExcelError("ensure-table needs --headers-json or --range")
    state = {}

    def build_plan(live_path):
        wb = _load(live_path)
        try:
            matches = _table_matches(wb, table_name)
            if len(matches) > 1:
                raise ExcelError(f"table name {table_name!r} is ambiguous")
            existing = matches[0] if matches else None
            ws = _sheet_by_name(wb, sheet)
            if ws is None and not create_sheet:
                raise ExcelError(
                    f"no sheet named {sheet!r}; pass --create-sheet to create it"
                )
            if ws is None and existing:
                raise ExcelError("existing table and requested new worksheet conflict")
            existing_table = existing[1] if existing else None
            if existing and existing[0].title.casefold() != sheet.casefold():
                raise ExcelError(
                    f"table {table_name!r} already exists on sheet {existing[0].title!r}"
                )

            if existing_table is not None:
                actual_headers = _table_headers(existing[0], existing_table)
                normalized_ref, bounds = _normalized_range(existing_table.ref)
                if cell_range:
                    requested_ref, _ = _normalized_range(cell_range)
                    if requested_ref != normalized_ref:
                        raise ExcelError(
                            "existing table range differs from the requested range"
                        )
                if requested_headers is not None and requested_headers != actual_headers:
                    raise ExcelError("existing table headers differ from the requested headers")
                metadata = _column_metadata(columns, actual_headers)
                changed = False
                if style:
                    current_style = getattr(
                        getattr(existing_table, "tableStyleInfo", None), "name", ""
                    )
                    changed = current_style != style
                for name, item in metadata.items():
                    col = bounds[0] + actual_headers.index(name)
                    if "number_format" in item:
                        changed = changed or any(
                            existing[0].cell(row_no, col).number_format != item["number_format"]
                            for row_no in range(bounds[1] + 1, bounds[3] + 1)
                        )
                    if "formula" in item:
                        origin = f"{get_column_letter(col)}{bounds[1] + 1}"
                        for row_no in range(bounds[1] + 1, bounds[3] + 1):
                            cell = existing[0].cell(row_no, col)
                            expected = Translator(
                                item["formula"], origin=origin
                            ).translate_formula(cell.coordinate)
                            if cell.value not in (None, expected):
                                raise ExcelError(
                                    f"formula metadata conflicts with {cell.coordinate}"
                                )
                            changed = changed or cell.value != expected
                state.update({
                    "action": "update", "sheet": existing[0].title,
                    "ref": normalized_ref, "bounds": bounds,
                    "headers": actual_headers, "metadata": metadata,
                    "range_digest": _range_digest(existing[0], bounds),
                })
                rows = max(0, bounds[3] - bounds[1] - _totals_count(existing_table))
                return office_tx.MutationPlan(
                    "ensure-table",
                    {"sheet": existing[0].title, "table": existing_table.displayName,
                     "range": normalized_ref, "ref": normalized_ref,
                     "row_count": rows, "headers": actual_headers,
                     "created_sheet": False},
                    {"affected": int(changed)}, changed=changed,
                )

            target_sheet = ws
            if cell_range:
                normalized_ref, bounds = _normalized_range(cell_range)
            else:
                normalized_ref, bounds = _normalized_range(
                    f"A1:{get_column_letter(len(requested_headers))}1"
                )
            width = bounds[2] - bounds[0] + 1
            if requested_headers is not None and len(requested_headers) != width:
                raise ExcelError("header count does not match the requested range width")
            if target_sheet is None:
                actual_headers = requested_headers
            else:
                for merged in target_sheet.merged_cells.ranges:
                    if _ranges_overlap(bounds, range_boundaries(str(merged))):
                        raise ExcelError("table range intersects merged cells")
                for other in target_sheet.tables.values():
                    if _ranges_overlap(bounds, range_boundaries(other.ref)):
                        raise ExcelError("table range overlaps an existing table")
                existing_headers = [
                    target_sheet.cell(bounds[1], col).value
                    for col in range(bounds[0], bounds[2] + 1)
                ]
                if requested_headers is None:
                    actual_headers = _validated_headers(existing_headers)
                else:
                    for current, requested in zip(existing_headers, requested_headers):
                        if current not in (None, requested):
                            raise ExcelError(
                                "requested headers would overwrite existing header values"
                            )
                    actual_headers = requested_headers
            metadata = _column_metadata(columns, actual_headers)
            if target_sheet is not None:
                # Validate formula conflicts without mutating the planning workbook.
                for name, item in metadata.items():
                    if "formula" not in item:
                        continue
                    col = bounds[0] + actual_headers.index(name)
                    origin = f"{get_column_letter(col)}{bounds[1] + 1}"
                    for row_no in range(bounds[1] + 1, bounds[3] + 1):
                        cell = target_sheet.cell(row_no, col)
                        expected = Translator(
                            item["formula"], origin=origin
                        ).translate_formula(cell.coordinate)
                        if cell.value not in (None, expected):
                            raise ExcelError(
                                f"formula metadata would overwrite populated cell {cell.coordinate}"
                            )
            state.update({
                "action": "create", "sheet": sheet, "ref": normalized_ref,
                "bounds": bounds, "headers": actual_headers,
                "metadata": metadata,
                "range_digest": _range_digest(target_sheet, bounds)
                    if target_sheet is not None else None,
            })
            rows = max(0, bounds[3] - bounds[1])
            return office_tx.MutationPlan(
                "ensure-table",
                {"sheet": sheet, "table": table_name, "range": normalized_ref,
                 "ref": normalized_ref, "row_count": rows,
                 "headers": actual_headers, "created_sheet": target_sheet is None},
                {"affected": 1},
            )
        finally:
            wb.close()

    def apply(stage, _plan):
        wb = _load(stage)
        try:
            ws = _sheet_by_name(wb, state["sheet"])
            if ws is None:
                ws = wb.create_sheet(state["sheet"])
            if state["action"] == "create":
                bounds = state["bounds"]
                for offset, header in enumerate(state["headers"]):
                    ws.cell(bounds[1], bounds[0] + offset).value = header
                try:
                    table = Table(displayName=table_name, ref=state["ref"])
                except ValueError as exc:
                    raise ExcelError(f"invalid Excel table name: {table_name!r}") from exc
                if style:
                    table.tableStyleInfo = TableStyleInfo(
                        name=style, showFirstColumn=False, showLastColumn=False,
                        showRowStripes=True, showColumnStripes=False,
                    )
                _apply_table_metadata(
                    ws, table, state["headers"], state["metadata"],
                    bounds, refuse_conflicts=True,
                )
                ws.add_table(table)
            else:
                _ws, table = _find_table(wb, table_name, state["sheet"])
                if style:
                    table.tableStyleInfo = TableStyleInfo(
                        name=style, showFirstColumn=False, showLastColumn=False,
                        showRowStripes=True, showColumnStripes=False,
                    )
                _apply_table_metadata(
                    ws, table, state["headers"], state["metadata"],
                    state["bounds"], refuse_conflicts=True,
                )
            wb.save(stage)
        finally:
            wb.close()

    def validate(stage, _plan):
        wb = _load(stage)
        try:
            ws, table = _find_table(wb, table_name, state["sheet"])
            if table.ref != state["ref"]:
                raise ExcelError("staged table range differs from the plan")
            if _table_headers(ws, table) != state["headers"]:
                raise ExcelError("staged table headers differ from the plan")
            if state["range_digest"] is not None and not state["metadata"]:
                if _range_digest(ws, state["bounds"]) != state["range_digest"]:
                    raise ExcelError("table conversion changed existing values or formulas")
            rows = max(0, state["bounds"][3] - state["bounds"][1]
                       - _totals_count(table))
            return {
                "sheet": ws.title, "table": table.displayName,
                "range": table.ref, "ref": table.ref, "row_count": rows,
            }
        finally:
            wb.close()

    try:
        return office_tx.execute_mutation(
            path, operation="ensure-table", plan=build_plan,
            apply=apply, validate=validate,
            expected_sha256=expected_sha256 or None, dry_run=dry_run,
        )
    except office_tx.TransactionError as exc:
        raise ExcelError(str(exc)) from exc


def append_table_row(
    path: str,
    table_name: str,
    row: dict,
    *,
    sheet: str = "",
    expected_sha256: str = "",
    dry_run: bool = False,
    coerce_iso_dates: bool = False,
    unique_columns: Optional[list[str]] = None,
) -> dict:
    if not isinstance(row, dict) or not row:
        raise ExcelError("row JSON must be a non-empty object")
    expected_ref = {}
    unique_columns = list(unique_columns or [])
    if len(unique_columns) != len(set(unique_columns)):
        raise ExcelError("unique columns must not contain duplicates")

    def build_plan(live_path):
        wb = _load(live_path)
        try:
            ws, table = _find_table(wb, table_name, sheet)
            headers = _table_headers(ws, table)
            unknown = [name for name in row if name not in headers]
            if unknown:
                raise ExcelError(f"unknown table column(s): {', '.join(unknown)}")
            unknown_unique = [name for name in unique_columns if name not in headers]
            if unknown_unique:
                raise ExcelError(
                    f"unknown unique column(s): {', '.join(unknown_unique)}"
                )
            missing_unique = [name for name in unique_columns if name not in row]
            if missing_unique:
                raise ExcelError(
                    f"row is missing unique column value(s): {', '.join(missing_unique)}"
                )
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            totals = _totals_count(table)
            data_end = max_row - totals
            target_row = max_row if totals else max_row + 1
            indexes = {name: min_col + headers.index(name) for name in headers}
            template_row = max(min_row, target_row - 1)
            coerced = {
                name: _coerce(
                    value,
                    date_column=_column_is_date(ws, template_row, indexes[name]),
                    coerce_iso_dates=coerce_iso_dates,
                )
                for name, value in row.items()
            }
            if unique_columns:
                matches = [
                    row_no for row_no in range(min_row + 1, data_end + 1)
                    if all(
                        _json_value(ws.cell(row_no, indexes[name]).value,
                                    ws.cell(row_no, indexes[name]))
                        == _json_value(coerced[name])
                        for name in unique_columns
                    )
                ]
                if matches:
                    if len(matches) == 1 and all(
                        _json_value(ws.cell(matches[0], indexes[name]).value,
                                    ws.cell(matches[0], indexes[name]))
                        == _json_value(value)
                        for name, value in coerced.items()
                    ):
                        expected_ref.update({
                            "old": table.ref, "new": table.ref,
                            "target_row": matches[0], "headers": headers,
                            "coerced": coerced, "idempotent": True,
                            "unique_columns": unique_columns,
                        })
                        return office_tx.MutationPlan(
                            "append-table-row",
                            {"table": table.displayName, "sheet": ws.title,
                             "range": table.ref, "ref": table.ref,
                             "row_count": max(0, data_end - min_row),
                             "appended": 0, "idempotent": True},
                            {"affected": 0}, changed=False,
                        )
                    raise ExcelConflict(
                        "table uniqueness constraint already exists",
                        {"table": table.displayName, "sheet": ws.title,
                         "columns": unique_columns, "match_count": len(matches)},
                    )
            for col in range(min_col, max_col + 1):
                if not totals and ws.cell(target_row, col).value is not None:
                    raise ExcelError("table expansion would overwrite populated cells")
            new_ref = (
                f"{get_column_letter(min_col)}{min_row}:"
                f"{get_column_letter(max_col)}{max_row + 1}"
            )
            expected_ref.update({
                "old": table.ref, "new": new_ref, "target_row": target_row,
                "headers": headers, "coerced": coerced, "idempotent": False,
                "unique_columns": unique_columns,
            })
            return office_tx.MutationPlan(
                "append-table-row",
                {"table": table.displayName, "sheet": ws.title,
                 "old_ref": table.ref, "new_ref": new_ref,
                 "range": new_ref, "ref": new_ref,
                 "affected_range": (
                     f"{get_column_letter(min_col)}{target_row}:"
                     f"{get_column_letter(max_col)}{target_row}"
                 ),
                 "row_count": max(0, data_end - min_row) + 1},
                {"affected": 1},
            )
        finally:
            wb.close()

    def apply(stage, _plan):
        wb = _load(stage)
        try:
            ws, table = _find_table(wb, table_name, sheet)
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            totals = _totals_count(table)
            target_row = max_row if totals else max_row + 1
            if totals:
                ws.insert_rows(target_row, 1)
            template_row = target_row - 1
            if template_row <= min_row:
                template_row = min_row
            if ws.row_dimensions[template_row].height is not None:
                ws.row_dimensions[target_row].height = ws.row_dimensions[template_row].height
            headers = expected_ref["headers"]
            for offset, header in enumerate(headers):
                col = min_col + offset
                source = ws.cell(template_row, col)
                target = ws.cell(target_row, col)
                _copy_cell_style(source, target)
                if header in row:
                    target.value = expected_ref["coerced"][header]
                elif isinstance(source.value, str) and source.value.startswith("="):
                    target.value = Translator(
                        source.value, origin=source.coordinate
                    ).translate_formula(target.coordinate)
                else:
                    calculated = _table_column_formula(table, offset)
                    target.value = calculated or None
            table.ref = expected_ref["new"]
            if getattr(table, "autoFilter", None) is not None:
                table.autoFilter.ref = expected_ref["new"]
            wb.save(stage)
        finally:
            wb.close()

    def validate(stage, _plan):
        wb = _load(stage)
        try:
            ws, table = _find_table(wb, table_name, sheet)
            if table.ref != expected_ref["new"]:
                raise ExcelError("staged table reference did not expand as planned")
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            data_end = max_row - _totals_count(table)
            if expected_ref["unique_columns"]:
                indexes = {
                    name: min_col + expected_ref["headers"].index(name)
                    for name in expected_ref["unique_columns"]
                }
                matches = [
                    row_no for row_no in range(min_row + 1, data_end + 1)
                    if all(
                        _json_value(ws.cell(row_no, indexes[name]).value,
                                    ws.cell(row_no, indexes[name]))
                        == _json_value(expected_ref["coerced"][name])
                        for name in expected_ref["unique_columns"]
                    )
                ]
                if len(matches) != 1:
                    raise ExcelError("staged uniqueness validation failed")
            return {
                "appended": 1, "range": table.ref, "ref": table.ref,
                "row_count": max(0, data_end - min_row),
            }
        finally:
            wb.close()

    try:
        return office_tx.execute_mutation(
            path, operation="append-table-row", plan=build_plan,
            apply=apply, validate=validate,
            expected_sha256=expected_sha256 or None, dry_run=dry_run,
        )
    except office_tx.TransactionError as exc:
        raise ExcelError(str(exc)) from exc


def update_table_row(
    path: str,
    table_name: str,
    key_column: str,
    key,
    updates: dict,
    *,
    sheet: str = "",
    expected_sha256: str = "",
    dry_run: bool = False,
    coerce_iso_dates: bool = False,
) -> dict:
    if not isinstance(updates, dict) or not updates:
        raise ExcelError("set JSON must be a non-empty object")
    target = {}

    def build_plan(live_path):
        wb = _load(live_path)
        try:
            ws, table = _find_table(wb, table_name, sheet)
            headers = _table_headers(ws, table)
            if key_column not in headers:
                raise ExcelError(f"unknown key column: {key_column}")
            unknown = [name for name in updates if name not in headers]
            if unknown:
                raise ExcelError(f"unknown table column(s): {', '.join(unknown)}")
            min_col, min_row, _, max_row = range_boundaries(table.ref)
            totals = _totals_count(table)
            data_end = max_row - totals
            key_col = min_col + headers.index(key_column)
            matches = [
                row_no for row_no in range(min_row + 1, data_end + 1)
                if ws.cell(row_no, key_col).value == key
                or (isinstance(key, str) and str(ws.cell(row_no, key_col).value) == key)
            ]
            if len(matches) != 1:
                raise ExcelError(
                    f"expected exactly one matching row; found {len(matches)}"
                )
            target.update({"row": matches[0], "min_col": min_col, "headers": headers})
            changed = sum(
                ws.cell(matches[0], min_col + headers.index(name)).value != value
                for name, value in updates.items()
            )
            return office_tx.MutationPlan(
                "update-table-row",
                {"table": table.displayName, "sheet": ws.title,
                 "matched": 1, "fields": len(updates),
                 "range": table.ref, "ref": table.ref,
                 "affected_range": (
                     f"{get_column_letter(min_col)}{matches[0]}:"
                     f"{get_column_letter(min_col + len(headers) - 1)}{matches[0]}"
                 ),
                 "row_count": max(0, data_end - min_row)},
                {"affected": changed},
                changed=bool(changed),
            )
        finally:
            wb.close()

    def apply(stage, _plan):
        wb = _load(stage)
        try:
            ws, _table = _find_table(wb, table_name, sheet)
            for name, value in updates.items():
                col = target["min_col"] + target["headers"].index(name)
                ws.cell(target["row"], col).value = _coerce(
                    value,
                    date_column=_column_is_date(ws, target["row"], col),
                    coerce_iso_dates=coerce_iso_dates,
                )
            wb.save(stage)
        finally:
            wb.close()

    def validate(stage, _plan):
        wb = _load(stage)
        try:
            ws, _table = _find_table(wb, table_name, sheet)
            for name, value in updates.items():
                col = target["min_col"] + target["headers"].index(name)
                expected = _coerce(
                    value,
                    date_column=_column_is_date(ws, target["row"], col),
                    coerce_iso_dates=coerce_iso_dates,
                )
                if ws.cell(target["row"], col).value != expected:
                    raise ExcelError(f"staged value validation failed for {name!r}")
            _ws, table = _find_table(wb, table_name, sheet)
            min_col, min_row, _, max_row = range_boundaries(table.ref)
            return {
                "updated": 1, "range": table.ref, "ref": table.ref,
                "row_count": max(0, max_row - _totals_count(table) - min_row),
            }
        finally:
            wb.close()

    try:
        return office_tx.execute_mutation(
            path, operation="update-table-row", plan=build_plan,
            apply=apply, validate=validate,
            expected_sha256=expected_sha256 or None, dry_run=dry_run,
        )
    except office_tx.TransactionError as exc:
        raise ExcelError(str(exc)) from exc
