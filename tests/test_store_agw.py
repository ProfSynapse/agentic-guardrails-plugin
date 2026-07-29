"""Archive store + agw CLI behavior: round-trips, undo, concurrency, conflicts."""
import ctypes
import json
import os
import subprocess
import sys
import threading
import time

import pytest

from core import store
from core import profiles
import scan_worker

REPO = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "plugin")
AGW = os.path.join(REPO, "scripts", "agw", "agw.py")


def run_agw(*args, env=None, check=True):
    e = dict(os.environ)
    if env:
        e.update(env)
    result = subprocess.run([sys.executable, AGW, *args],
                            capture_output=True, text=True, encoding="utf-8", env=e)
    if check and result.returncode != 0:
        raise AssertionError(f"agw {' '.join(args)} failed: {result.stderr}")
    return result


def test_archive_restore_roundtrip(tmp_path):
    f = tmp_path / "report.txt"
    f.write_text("important content")
    entry = store.archive_file(str(f), mode="move", reason="test")
    assert not f.exists()
    assert os.path.exists(entry["dest"])
    store.restore(str(f))
    assert f.read_text() == "important content"


def test_versions_monotonic(tmp_path):
    f = tmp_path / "doc.txt"
    for i in range(3):
        f.write_text(f"version {i}")
        store.archive_file(str(f), mode="copy")
    versions = [e["version"] for e in store.list_versions(str(f))]
    assert versions == [1, 2, 3]


def test_restore_specific_version(tmp_path):
    f = tmp_path / "doc.txt"
    for i in range(3):
        f.write_text(f"version {i}")
        store.archive_file(str(f), mode="copy")
    store.restore(str(f), version=1)
    assert f.read_text() == "version 0"


def test_restore_never_clobbers_silently(tmp_path):
    f = tmp_path / "doc.txt"
    f.write_text("old")
    store.archive_file(str(f), mode="copy")
    f.write_text("newer work")
    store.restore(str(f))
    # the "newer work" must itself have been archived before restore
    contents = [open(e["dest"]).read() for e in store.list_versions(str(f))
                if os.path.isfile(e["dest"])]
    assert "newer work" in contents


def test_undo_archive(tmp_path):
    f = tmp_path / "x.txt"
    f.write_text("data")
    store.archive_file(str(f), mode="move")
    assert not f.exists()
    store.undo_last()
    assert f.read_text() == "data"


def test_undo_move(tmp_path):
    src = tmp_path / "a.txt"
    src.write_text("1")
    dest = tmp_path / "sub" / "b.txt"
    store.logged_move(str(src), str(dest))
    assert dest.exists() and not src.exists()
    store.undo_last()
    assert src.exists() and not dest.exists()


def test_pre_image_dedupe(tmp_path):
    f = tmp_path / "doc.txt"
    f.write_text("same content")
    e1 = store.archive_file(str(f), mode="copy", dedupe=True)
    e2 = store.archive_file(str(f), mode="copy", dedupe=True)
    assert e1["version"] == 1
    assert e2.get("deduped") is True


def test_concurrent_archives_no_lost_versions(tmp_path):
    files = []
    for i in range(12):
        f = tmp_path / f"f{i}.txt"
        f.write_text(f"content {i}")
        files.append(str(f))
    errors = []

    def worker(path):
        try:
            store.archive_file(path, mode="move")
        except Exception as exc:  # noqa: BLE001
            errors.append(exc)

    threads = [threading.Thread(target=worker, args=(p,)) for p in files]
    for t in threads:
        t.start()
    for t in threads:
        t.join()
    assert not errors
    ops = [o for o in store.oplog_read() if o.get("op") == "archive"]
    assert len(ops) == 12


def test_path_with_spaces_and_unicode(tmp_path):
    f = tmp_path / "Q3 report — final (v2).txt"
    f.write_text("data")
    store.archive_file(str(f), mode="move")
    store.restore(str(f))
    assert f.read_text() == "data"


# --- CLI-level tests ----------------------------------------------------------

def test_cli_checkout_publish_roundtrip(tmp_path, agw_home):
    f = tmp_path / "notes.txt"
    f.write_text("original text")
    run_agw("checkout", str(f))
    working = tmp_path / "_workspace" / "notes.txt"
    assert working.read_text() == "original text"
    working.write_text("edited text")
    run_agw("publish", str(f))
    assert f.read_text() == "edited text"
    # prior version archived
    assert any(e["op"] == "archive" for e in store.oplog_read())
    run_agw("restore", str(f))
    assert f.read_text() == "original text"


def test_cli_publish_conflict_detection(tmp_path, agw_home):
    f = tmp_path / "doc.txt"
    f.write_text("base")
    run_agw("checkout", str(f))
    f.write_text("someone else edited this")  # simulate external edit
    (tmp_path / "_workspace" / "doc.txt").write_text("agent edit")
    result = run_agw("publish", str(f), check=False)
    assert result.returncode == 3
    assert "CONFLICT" in result.stderr
    assert f.read_text() == "someone else edited this"  # live file untouched


def test_xlsx_checkout_defaults_to_style_preserving_workbook(tmp_path, agw_home):
    openpyxl = pytest.importorskip("openpyxl")
    from openpyxl.styles import PatternFill

    source = tmp_path / "dashboard.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Dashboard"
    sheet["A1"] = "Metric"
    sheet["A1"].fill = PatternFill("solid", fgColor="00FF00")
    sheet["B2"] = "=1+1"
    workbook.save(source)

    checkout = run_agw("checkout", str(source), "--json")
    checkout_data = json.loads(checkout.stdout)
    working = tmp_path / "_workspace" / "dashboard.xlsx"
    assert checkout_data["checkout_mode"] == "preserve"
    assert checkout_data["lossy"] is False
    assert working.exists()
    assert not list((tmp_path / "_workspace").glob("*.csv"))

    edited = openpyxl.load_workbook(working)
    edited["Dashboard"]["C2"] = "new"
    edited.save(working)
    edited.close()
    published = run_agw("publish", str(source), "--json")
    assert json.loads(published.stdout)["checkout_mode"] == "preserve"

    final = openpyxl.load_workbook(source, data_only=False)
    assert final["Dashboard"]["B2"].value == "=1+1"
    assert final["Dashboard"]["A1"].fill.fgColor.rgb.endswith("00FF00")
    assert final["Dashboard"]["C2"].value == "new"
    final.close()


def test_cli_scan_reports_stubs(tmp_path, agw_home):
    (tmp_path / "Budget.gsheet").write_text(json.dumps({"url": "x"}))
    (tmp_path / "real.txt").write_text("hello")
    result = run_agw("scan", str(tmp_path), "--json")
    data = json.loads(result.stdout)
    assert "Budget.gsheet" in data["gdoc_stubs"]
    assert data["files"] == 2
    assert data["complete"] is True


def test_cli_scan_rejects_file_path(tmp_path, agw_home):
    target = tmp_path / "not-a-directory.txt"
    target.write_text("content")
    result = run_agw("scan", str(target), "--json", check=False)
    assert result.returncode == 2
    assert "requires a directory" in result.stderr


def test_cli_scan_bounds_return_partial_results(tmp_path, agw_home):
    for index in range(8):
        (tmp_path / f"item-{index}.txt").write_text(str(index))
    result = run_agw(
        "scan", str(tmp_path), "--max-files", "3", "--no-size", "--json"
    )
    data = json.loads(result.stdout)
    assert data["complete"] is False
    assert data["stop_reason"] == "max_files"
    assert data["files_inspected"] == 3
    assert data["bytes"] is None
    assert data["elapsed_seconds"] >= 0
    assert data["placeholder_detection"] == "limited"
    assert data["deadline_enforced"] is True
    assert data["worker_cleanup"] == "complete"


def _scan_request(path, *, seconds=0.6, no_size=True, block=""):
    started = time.monotonic()
    return {
        "path": str(path),
        "started_at": started,
        "worker_deadline": started + seconds,
        "max_seconds": seconds,
        "max_files": 100,
        "max_depth": 4,
        "no_size": no_size,
        "profile_override": "auto",
        "_test_block_stage": block,
        "_test_block_seconds": 30.0,
    }


def _pid_exists(pid):
    if not pid:
        return False
    if os.name == "nt":
        process_query_limited_information = 0x1000
        handle = ctypes.windll.kernel32.OpenProcess(
            process_query_limited_information, False, pid
        )
        if not handle:
            return False
        exit_code = ctypes.c_ulong()
        ctypes.windll.kernel32.GetExitCodeProcess(handle, ctypes.byref(exit_code))
        ctypes.windll.kernel32.CloseHandle(handle)
        return exit_code.value == 259  # STILL_ACTIVE
    try:
        os.kill(pid, 0)
    except OSError:
        return False
    return True


@pytest.mark.parametrize("stage,no_size", [
    ("path_validation", True),
    ("profile_detection", True),
    ("scandir", True),
    ("scandir_next", True),
    ("is_dir", True),
    ("stat", False),
])
def test_scan_blocking_filesystem_stages_obey_parent_deadline(
        tmp_path, monkeypatch, stage, no_size):
    (tmp_path / "item.txt").write_text("metadata")
    monkeypatch.setenv("AGW_TEST_MODE", "1")
    request = _scan_request(tmp_path, seconds=0.6, no_size=no_size, block=stage)
    wall_start = time.monotonic()
    result, fatal = scan_worker.run_bounded_scan(request)
    wall_elapsed = time.monotonic() - wall_start

    assert fatal is None
    assert result["complete"] is False
    assert result["stop_reason"] == "max_seconds"
    assert result["elapsed_seconds"] <= 0.6
    assert wall_elapsed < 1.0
    assert result["worker_terminated"] is True
    assert result["worker_cleanup"] == "complete"
    assert not _pid_exists(result["_worker_pid"])
    assert json.loads(json.dumps(result))["stop_reason"] == "max_seconds"
    if stage == "profile_detection":
        assert result["profile_detection"] == "timed_out"
    if stage in {"scandir", "scandir_next", "is_dir", "stat"}:
        assert result["directories_inspected"] == 1
    if stage in {"is_dir", "stat"}:
        assert result["entries_seen"] >= 1
    if stage == "stat":
        assert result["files_inspected"] == 1


def test_scan_max_depth_stops_before_descending(tmp_path):
    child = tmp_path / "child"
    child.mkdir()
    (child / "nested.txt").write_text("nested")
    result = run_agw(
        "scan", str(tmp_path), "--max-depth", "0", "--no-size", "--json"
    )
    data = json.loads(result.stdout)
    assert data["complete"] is False
    assert data["stop_reason"] == "max_depth"
    assert data["directories_inspected"] == 1
    assert data["files_inspected"] == 0


def test_scan_local_folder_completes_quickly(tmp_path):
    (tmp_path / "one.txt").write_text("one")
    (tmp_path / "two.json").write_text("{}")
    started = time.monotonic()
    result = run_agw(
        "scan", str(tmp_path), "--max-seconds", "5", "--no-size", "--json"
    )
    wall_elapsed = time.monotonic() - started
    data = json.loads(result.stdout)
    assert data["complete"] is True
    assert data["files_inspected"] == 2
    assert wall_elapsed < 2.0


def test_scan_emoji_path_overrides_legacy_worker_encoding(tmp_path):
    folder = tmp_path / "Clients 👥"
    folder.mkdir()
    (folder / "record.txt").write_text("metadata", encoding="utf-8")
    result = run_agw(
        "scan", str(folder), "--fast", "--json",
        env={"PYTHONIOENCODING": "cp1252"}, check=False,
    )
    assert result.returncode == 0, result.stderr
    data = json.loads(result.stdout)
    assert data["complete"] is True
    assert data["files_inspected"] == 1
    assert data["path"] == str(folder)
    assert data["worker_cleanup"] == "complete"


def test_worker_json_emission_has_ascii_escape_fallback(monkeypatch):
    class LegacyStream:
        def __init__(self):
            self.writes = []

        def write(self, value):
            if any(ord(char) > 127 for char in value):
                raise UnicodeEncodeError("cp1252", value, 0, 1, "unsupported")
            self.writes.append(value)

        def flush(self):
            pass

    stream = LegacyStream()
    monkeypatch.setattr(scan_worker.sys, "stdout", stream)
    scan_worker._write_json_line({"path": "Clients 👥"})
    assert len(stream.writes) == 1
    assert "\\ud83d\\udc65" in stream.writes[0]
    assert json.loads(stream.writes[0])["path"] == "Clients 👥"


def test_unexplained_worker_exit_includes_bounded_stderr_tail(
        tmp_path, monkeypatch):
    monkeypatch.setenv("AGW_TEST_MODE", "1")
    request = _scan_request(tmp_path, seconds=2.0, no_size=True)
    request["_test_worker_exit"] = 17
    request["_test_worker_stderr"] = \
        "discarded-prefix-" + ("x" * 100_000) + " actionable-tail 👥"
    result, fatal = scan_worker.run_bounded_scan(request)

    assert fatal is None
    assert result["complete"] is False
    assert result["stop_reason"] == "worker_error"
    assert result["worker_cleanup"] == "complete"
    assert not _pid_exists(result["_worker_pid"])
    error = result["errors"][0]
    assert error["error"] == "WorkerExited"
    assert error["returncode"] == 17
    assert error["detail"].endswith("actionable-tail 👥")
    assert len(error["detail"]) <= scan_worker.MAX_STDERR_TAIL_CHARS
    assert "discarded-prefix" not in error["detail"]


def test_scan_filesystem_error_is_structured_json(tmp_path):
    missing = tmp_path / "missing"
    # Repeat because short-lived Windows subprocesses previously raced the
    # stdout reader and lost their final fatal frame under suite load.
    for _ in range(5):
        result = run_agw("scan", str(missing), "--json", check=False)
        assert result.returncode == 2
        data = json.loads(result.stderr)
        assert data["ok"] is False
        assert data["error"]["code"] == "path_not_found"


def test_sync_profile_precedes_enclosing_git_marker(tmp_path):
    sync_root = tmp_path / "shared"
    sync_root.mkdir()
    (sync_root / ".dropbox").mkdir()
    repo = sync_root / "repo"
    repo.mkdir()
    (repo / ".git").mkdir()
    profiles._cache.clear()
    assert profiles.detect(str(repo)).name == "dropbox"


def test_google_drive_shared_drives_mount_is_classified(tmp_path):
    mounted = tmp_path / "Shared drives" / "Synaptic Labs"
    mounted.mkdir(parents=True)
    profiles._cache.clear()
    assert profiles.detect(str(mounted), assume_directory=True).name == "gdrive-sync"


def test_provider_volume_label_is_classified_without_registry_calls(
        tmp_path, monkeypatch):
    mounted = tmp_path / "mounted" / "team"
    mounted.mkdir(parents=True)
    monkeypatch.setattr(profiles, "_windows_volume_label",
                        lambda _path: "Google Drive")
    profiles._cache.clear()
    assert profiles.detect(str(mounted), assume_directory=True).name == "gdrive-sync"


def test_profile_detection_does_not_enumerate_ancestors(tmp_path, monkeypatch):
    root = tmp_path / "sync"
    nested = root / "project"
    nested.mkdir(parents=True)
    (root / ".dropbox").mkdir()
    monkeypatch.setattr(profiles.os, "listdir", lambda _path: (_ for _ in ()).throw(
        AssertionError("profile detection enumerated an ancestor")
    ))
    profiles._cache.clear()
    assert profiles.detect(str(nested), assume_directory=True).name == "dropbox"


def test_fast_no_size_scan_avoids_content_reads_and_stat(
        tmp_path, monkeypatch):
    folder = tmp_path / "OneDrive - Example"
    folder.mkdir()
    target = folder / "cloud.dat"
    target.write_bytes(b"metadata-only")
    calls = {"stat": 0}

    class Entry:
        name = target.name
        path = str(target)

        def is_dir(self, *, follow_symlinks=False):
            assert follow_symlinks is False
            return False

        def stat(self, *, follow_symlinks=False):
            calls["stat"] += 1
            raise AssertionError("fast/no-size scan called entry.stat")

    class Iterator:
        def __init__(self):
            self.items = iter([Entry()])

        def __next__(self):
            return next(self.items)

        def close(self):
            pass

    monkeypatch.setattr(scan_worker.os, "scandir", lambda _path: Iterator())
    monkeypatch.setattr("builtins.open", lambda *a, **k: (_ for _ in ()).throw(
        AssertionError("scan opened file content")
    ))
    request = _scan_request(folder, seconds=5.0, no_size=True)
    request["profile_override"] = "gdrive-sync"
    data, fatal = scan_worker.scan_in_process(request)
    assert fatal is None
    assert data["files_inspected"] == 1
    assert data["placeholder_detection"] == "limited"
    assert data["placeholders"] == []
    assert calls["stat"] == 0


def test_cli_snapshot_preflight(tmp_path, agw_home):
    big = tmp_path / "big.bin"
    big.write_bytes(b"x" * 1000)
    result = run_agw("snapshot", str(tmp_path), check=False,
                     env={"AGW_SNAPSHOT_MAX_BYTES": "100"})
    assert result.returncode == 3
    assert "force" in result.stderr.lower()
    # and --force overrides
    result = run_agw("snapshot", str(tmp_path), "--force",
                     env={"AGW_SNAPSHOT_MAX_BYTES": "100"})
    assert result.returncode == 0


def test_cli_prune_refuses_without_human_flag(agw_home):
    result = run_agw("prune", check=False)
    assert result.returncode == 4
    assert "human" in result.stderr.lower() or "refusing" in result.stderr.lower()


def test_cli_doctor_runs(agw_home):
    result = run_agw("doctor", "--json")
    data = json.loads(result.stdout)
    assert data["agw_home_writable"] is True


def test_cli_doctor_uses_real_archive_write_probe(tmp_path):
    blocked_home = tmp_path / "blocked-home"
    blocked_home.mkdir()
    (blocked_home / "archive").write_text("not a directory")
    result = run_agw("doctor", "--json", env={"AGW_HOME": str(blocked_home)})
    data = json.loads(result.stdout)
    assert data["agw_home_writable"] is False


def test_cli_archive_permission_failure_is_plain_and_preserves_source(tmp_path):
    blocked_home = tmp_path / "blocked-home"
    blocked_home.mkdir()
    (blocked_home / "archive").write_text("not a directory")
    source = tmp_path / "keep-me.txt"
    source.write_text("still here")

    result = run_agw(
        "archive", str(source), check=False, env={"AGW_HOME": str(blocked_home)}
    )

    assert result.returncode != 0
    assert source.read_text() == "still here"
    assert "original file was not moved or changed" in result.stderr.lower()
    assert "host's normal approval" in result.stderr.lower()
    assert "traceback" not in result.stderr.lower()
    assert str(blocked_home).lower() not in result.stderr.lower()
