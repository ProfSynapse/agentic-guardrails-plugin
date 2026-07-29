"""Structured Office reads and guarded mutation contracts."""
import datetime as dt
import json
import os
from pathlib import Path
import shutil
import subprocess
import sys
import zipfile

import pytest

openpyxl = pytest.importorskip("openpyxl")
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

REPO = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "plugin")
AGW = os.path.join(REPO, "scripts", "agw", "agw.py")
sys.path.insert(0, os.path.join(REPO, "scripts", "agw"))

import office_excel  # noqa: E402
import office  # noqa: E402
import office_tx  # noqa: E402
import office_word  # noqa: E402
from core import store  # noqa: E402


def run_agw(*args, env=None, check=True, input_text=None):
    process_env = dict(os.environ)
    if env:
        process_env.update(env)
    result = subprocess.run(
        [sys.executable, AGW, *args],
        capture_output=True, text=True, encoding="utf-8",
        env=process_env, input=input_text,
    )
    if check and result.returncode:
        raise AssertionError(result.stderr)
    return result


@pytest.fixture
def table_book(tmp_path):
    path = tmp_path / "table.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["ID", "Status", "Due", "Total"])
    ws.append(["A-1", "Open", dt.date(2026, 7, 28), "=1+1"])
    ws["A997"].fill = PatternFill("solid", fgColor="FF0000")
    table = Table(displayName="Orders", ref="A1:D2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True
    )
    ws.add_table(table)
    wb.save(path)
    return str(path)


@pytest.fixture
def empty_book(tmp_path):
    path = tmp_path / "empty.xlsx"
    openpyxl.Workbook().save(path)
    return str(path)


def _inject_lossy_extension(path):
    rewritten = str(path) + ".rewritten"
    with zipfile.ZipFile(path) as source, zipfile.ZipFile(
            rewritten, "w", zipfile.ZIP_DEFLATED) as target:
        for item in source.infolist():
            payload = source.read(item)
            if item.filename == "xl/worksheets/sheet1.xml":
                extension = (
                    b'<extLst><ext uri="{synthetic-preservation-risk}">'
                    b'<x14:conditionalFormattings '
                    b'xmlns:x14="http://schemas.microsoft.com/office/'
                    b'spreadsheetml/2009/9/main"/></ext></extLst>'
                )
                payload = payload.replace(b"</worksheet>", extension + b"</worksheet>")
            target.writestr(item, payload)
    os.replace(rewritten, path)


def _inject_x14_x15_extensions(path):
    rewritten = str(path) + ".x14-x15"
    with zipfile.ZipFile(path) as source, zipfile.ZipFile(
            rewritten, "w", zipfile.ZIP_DEFLATED) as target:
        for item in source.infolist():
            payload = source.read(item)
            if item.filename == "xl/worksheets/sheet1.xml":
                extension = (
                    b'<extLst><ext uri="{x14-x15-preservation-risk}">'
                    b'<x14:conditionalFormattings '
                    b'xmlns:x14="http://schemas.microsoft.com/office/'
                    b'spreadsheetml/2009/9/main"/>'
                    b'<x15:conditionalFormattings '
                    b'xmlns:x15="http://schemas.microsoft.com/office/'
                    b'spreadsheetml/2010/11/main"/>'
                    b'</ext></extLst>'
                )
                payload = payload.replace(b"</worksheet>", extension + b"</worksheet>")
            target.writestr(item, payload)
    os.replace(rewritten, path)


def _inject_safe_style_metadata(path):
    rewritten = str(path) + ".safe-metadata"
    with zipfile.ZipFile(path) as source, zipfile.ZipFile(
            rewritten, "w", zipfile.ZIP_DEFLATED) as target:
        for item in source.infolist():
            payload = source.read(item)
            if item.filename == "xl/styles.xml":
                payload = payload.replace(
                    b"<fonts count=",
                    (b'<fonts xmlns:x14ac="http://schemas.microsoft.com/office/'
                     b'spreadsheetml/2009/9/ac" x14ac:knownFonts="1" count='),
                    1,
                )
            target.writestr(item, payload)
    os.replace(rewritten, path)


def test_info_uses_actual_values_not_styled_extent(table_book):
    data = office_excel.workbook_info(table_book)
    sheet = data["sheets"][0]
    assert sheet["used_range"] == "A1:D2"
    assert sheet["rows"] == 2
    assert sheet["worksheet_dimension"] == "A1:D997"
    assert sheet["tables"][0]["headers"] == ["ID", "Status", "Due", "Total"]


def test_read_table_is_typed_and_headers_once(table_book):
    data = office_excel.read_table(
        table_book, "Orders", columns=["ID", "Due"], where={"Status": "Open"}
    )
    assert data["headers"] == ["ID", "Due"]
    assert data["rows"] == [["A-1", "2026-07-28"]]
    assert len(data["hash"]) == 64
    assert data["preservation"]["safe_to_mutate"] is True


def test_ensure_table_creates_sheet_and_is_idempotent(empty_book, agw_home):
    before = store.file_sha256(empty_book)
    result = office_excel.ensure_table(
        empty_book, "RecordsTable", sheet="Records",
        headers=["RecordID", "Status", "Amount"], create_sheet=True,
        style="TableStyleMedium2",
        columns={"Amount": {"number_format": "0.00"}},
        expected_sha256=before,
    )
    assert result["changed"] == 1
    assert result["before_hash"] == before
    assert result["after_hash"] == store.file_sha256(empty_book)
    assert result["sheet"] == "Records"
    assert result["table"] == "RecordsTable"
    assert result["range"] == "A1:C1"
    wb = openpyxl.load_workbook(empty_book)
    assert wb["Records"].tables["RecordsTable"].ref == "A1:C1"
    assert [wb["Records"].cell(1, col).value for col in range(1, 4)] == [
        "RecordID", "Status", "Amount",
    ]
    wb.close()

    after_first = store.file_sha256(empty_book)
    repeat = office_excel.ensure_table(
        empty_book, "RecordsTable", sheet="Records",
        headers=["RecordID", "Status", "Amount"], create_sheet=True,
        style="TableStyleMedium2",
        columns={"Amount": {"number_format": "0.00"}},
        expected_sha256=after_first,
    )
    assert repeat["changed"] == 0
    assert repeat["before_hash"] == repeat["after_hash"] == after_first
    assert len(store.list_versions(empty_book)) == 1


def test_ensure_table_converts_explicit_range_without_changing_values(tmp_path):
    path = tmp_path / "range.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Records"
    ws.append(["RecordID", "Status", "Total"])
    ws.append(["R-1", "Open", "=1+1"])
    ws.append(["R-2", "Closed", "=2+2"])
    wb.save(path)
    before_values = [
        [ws.cell(row, col).value for col in range(1, 4)] for row in range(1, 4)
    ]
    result = office_excel.ensure_table(
        str(path), "RecordsTable", sheet="Records", cell_range="A1:C3"
    )
    assert result["changed"] == 1
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb["Records"]
    assert ws.tables["RecordsTable"].ref == "A1:C3"
    assert [
        [ws.cell(row, col).value for col in range(1, 4)] for row in range(1, 4)
    ] == before_values
    wb.close()


@pytest.mark.parametrize("headers,cell_range,message", [
    (["RecordID", "Status"], "A1:C2", "header count"),
    (["RecordID", "Other", "Total"], "A1:C2", "overlap"),
    (["RecordID", "recordid", "Total"], "", "unique"),
])
def test_ensure_table_refuses_malformed_or_destructive_requests(
        table_book, headers, cell_range, message):
    before = store.file_sha256(table_book)
    with pytest.raises(office_excel.ExcelError, match=message):
        office_excel.ensure_table(
            table_book, "NewTable", sheet="Data", headers=headers,
            cell_range=cell_range,
        )
    assert store.file_sha256(table_book) == before


def test_append_expands_table_copies_style_and_formula(table_book, agw_home):
    before = store.file_sha256(table_book)
    result = office_excel.append_table_row(
        table_book, "Orders",
        {"ID": "A-2", "Status": "Closed", "Due": "2026-08-01"},
        expected_sha256=before,
    )
    assert result["appended"] == 1
    assert result["snapshot"] == 1
    wb = openpyxl.load_workbook(table_book)
    ws = wb["Data"]
    assert ws.tables["Orders"].ref == "A1:D3"
    assert ws["A3"].value == "A-2"
    assert ws["C3"].value == dt.datetime(2026, 8, 1)
    assert ws["D3"].value == "=1+1"
    assert ws["A3"].style_id == ws["A2"].style_id
    versions = store.list_versions(table_book)
    assert len(versions) == 1
    assert versions[0]["sha256"] == before


def test_append_translates_relative_formulas_and_preserves_table_settings(
        tmp_path, agw_home):
    path = tmp_path / "calculated.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Records"
    ws.append(["RecordID", "Amount", "Double"])
    ws.append(["R-1", 4, "=B2*2"])
    ws.append(["", "Total", "=SUM(C2:C2)"])
    table = Table(displayName="RecordsTable", ref="A1:C3")
    table.totalsRowCount = 1
    table.totalsRowShown = True
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4", showRowStripes=True, showFirstColumn=True,
    )
    ws.add_table(table)
    wb.save(path)

    result = office_excel.append_table_row(
        str(path), "RecordsTable", {"RecordID": "R-2", "Amount": 7}
    )
    assert result["range"] == "A1:C4"
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb["Records"]
    table = ws.tables["RecordsTable"]
    assert table.ref == "A1:C4"
    assert table.autoFilter.ref == "A1:C4"
    assert table.totalsRowCount == 1
    assert table.tableStyleInfo.name == "TableStyleMedium4"
    assert table.tableStyleInfo.showFirstColumn is True
    assert ws["C3"].value == "=B3*2"
    assert ws["C4"].value == "=SUM(C2:C2)"
    assert ws["B3"].number_format == ws["B2"].number_format
    wb.close()


@pytest.mark.parametrize("value", [
    "=WEBSERVICE(\"https://example.invalid\")",
    {"$formula": "='[external.xlsx]Sheet1'!A1"},
])
def test_table_mutations_reject_unsafe_or_untyped_formulas(
        table_book, value, agw_home):
    before = store.file_sha256(table_book)
    with pytest.raises(office_excel.ExcelError, match="formula|external"):
        office_excel.append_table_row(
            table_book, "Orders", {"ID": "A-2", "Total": value}
        )
    assert store.file_sha256(table_book) == before
    assert store.list_versions(table_book) == []


def test_append_single_column_uniqueness_is_atomic_and_idempotent(
        table_book, agw_home):
    before = store.file_sha256(table_book)
    retry = office_excel.append_table_row(
        table_book, "Orders", {"ID": "A-1", "Status": "Open"},
        unique_columns=["ID"], expected_sha256=before,
    )
    assert retry["changed"] == 0
    assert retry["idempotent"] is True
    assert store.list_versions(table_book) == []

    with pytest.raises(office_excel.ExcelConflict) as caught:
        office_excel.append_table_row(
            table_book, "Orders", {"ID": "A-1", "Status": "Closed"},
            unique_columns=["ID"], expected_sha256=before,
        )
    assert caught.value.error_code == "uniqueness_conflict"
    assert caught.value.details["columns"] == ["ID"]
    assert store.file_sha256(table_book) == before
    assert store.list_versions(table_book) == []


def test_append_composite_uniqueness_conflict_is_atomic(table_book, agw_home):
    before = store.file_sha256(table_book)
    with pytest.raises(office_excel.ExcelConflict) as caught:
        office_excel.append_table_row(
            table_book, "Orders", {"ID": "A-1", "Status": "Open", "Due": "later"},
            unique_columns=["ID", "Status"], expected_sha256=before,
        )
    assert caught.value.details["columns"] == ["ID", "Status"]
    assert store.file_sha256(table_book) == before
    assert store.list_versions(table_book) == []


def test_update_dry_run_has_no_archive_then_commits(table_book, agw_home):
    before = store.file_sha256(table_book)
    dry = office_excel.update_table_row(
        table_book, "Orders", "ID", "A-1", {"Status": "Closed"},
        expected_sha256=before, dry_run=True,
    )
    assert dry["dry_run"] is True
    assert store.list_versions(table_book) == []
    assert store.file_sha256(table_book) == before

    result = office_excel.update_table_row(
        table_book, "Orders", "ID", "A-1", {"Status": "Closed"},
        expected_sha256=before,
    )
    assert result["updated"] == 1
    wb = openpyxl.load_workbook(table_book)
    assert wb["Data"]["B2"].value == "Closed"


def test_hash_conflict_refuses_without_snapshot(table_book, agw_home):
    with pytest.raises(office_excel.ExcelError, match="CONFLICT"):
        office_excel.update_table_row(
            table_book, "Orders", "ID", "A-1", {"Status": "Closed"},
            expected_sha256="0" * 64,
        )
    assert store.list_versions(table_book) == []


def test_lossy_ooxml_extension_is_reported_and_mutation_is_refused(
        table_book, agw_home):
    _inject_lossy_extension(table_book)
    before = store.file_sha256(table_book)
    info = office_excel.workbook_info(table_book)
    assert info["preservation"]["safe_to_mutate"] is False
    assert info["preservation"]["risks"][0]["part"] == \
        "xl/worksheets/sheet1.xml"
    assert info["preservation"]["risks"][0]["extension_uri"] == \
        "{synthetic-preservation-risk}"
    assert info["preservation"]["risks"][0]["namespaces"] == [
        "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
    ]

    with pytest.raises(office_excel.ExcelError, match="preservation"):
        office_excel.append_table_row(
            table_book, "Orders", {"ID": "A-2"}, expected_sha256=before,
        )
    assert store.file_sha256(table_book) == before
    assert store.list_versions(table_book) == []
    assert office_tx.transaction_status() == []
    assert not list(Path(table_book).parent.glob(".agw-office-*"))


def test_set_cell_surgically_preserves_unknown_excel_extensions(
        table_book, agw_home):
    _inject_x14_x15_extensions(table_book)
    with zipfile.ZipFile(table_book) as package:
        before = {item.filename: package.read(item) for item in package.infolist()}
    expected = store.file_sha256(table_book)

    result = office.set_cell(
        table_book, "Data", "B2", "57", expected_sha256=expected,
    )
    assert result["adapter"] == "ooxml-surgical"
    assert result["preservation"]["unknown_parts_preserved"] is True
    assert result["old"] == "Open"
    assert result["new"] == 57

    with zipfile.ZipFile(table_book) as package:
        after = {item.filename: package.read(item) for item in package.infolist()}
    assert before.keys() == after.keys()
    worksheet = "xl/worksheets/sheet1.xml"
    assert before[worksheet] != after[worksheet]
    for name in before:
        if name != worksheet:
            assert before[name] == after[name], name
    before_extension = before[worksheet][before[worksheet].index(b"<extLst>"):]
    after_extension = after[worksheet][after[worksheet].index(b"<extLst>"):]
    assert before_extension == after_extension
    with pytest.warns(UserWarning, match="extension"):
        workbook = openpyxl.load_workbook(table_book, data_only=False)
    assert workbook["Data"]["B2"].value == 57
    workbook.close()
    assert len(store.list_versions(table_book)) == 1
    assert office_tx.transaction_status() == []


def test_known_non_destructive_style_metadata_is_allowlisted(
        table_book, agw_home):
    _inject_safe_style_metadata(table_book)
    risks = office_tx.inspect_preservation_risks(table_book)
    assert risks == []
    info = office_excel.workbook_info(table_book)
    assert info["preservation"]["safe_to_mutate"] is True


def test_normalize_removes_only_allowlisted_metadata_to_guarded_output(
        table_book, tmp_path, agw_home):
    _inject_safe_style_metadata(table_book)
    output = tmp_path / "normalized.xlsx"
    result = office_tx.normalize_safe_metadata(
        table_book, str(output), expected_sha256=store.file_sha256(table_book),
        expected_output_sha256="absent",
    )
    assert result["removed_count"] == 1
    assert result["snapshot_state"] == "ABSENT"
    assert output.exists()
    assert office_tx.inspect_preservation_risks(str(output)) == []
    with zipfile.ZipFile(output) as package:
        styles = package.read("xl/styles.xml")
    assert b"knownFonts" not in styles
    assert not list(tmp_path.glob(".agw-office-normalize-*"))


def test_normalize_refuses_unknown_extension_without_output(
        table_book, tmp_path, agw_home):
    _inject_lossy_extension(table_book)
    output = tmp_path / "normalized.xlsx"
    with pytest.raises(office_tx.PreservationError):
        office_tx.normalize_safe_metadata(
            table_book, str(output), expected_output_sha256="absent",
        )
    assert not output.exists()
    assert store.discover_archive_transactions() == []


def test_cli_read_table_compact_json(table_book, agw_home):
    result = run_agw(
        "office", "read-table", table_book, "--table", "Orders",
        "--columns", "ID,Status", "--limit", "1", "--json",
    )
    assert "\n" not in result.stdout.rstrip("\n")
    data = json.loads(result.stdout)
    assert data["headers"] == ["ID", "Status"]
    assert data["rows"] == [["A-1", "Open"]]


def test_cli_read_table_returns_cached_value_and_formula(table_book, agw_home):
    result = run_agw(
        "office", "read-table", table_book, "--table", "Orders",
        "--columns", "ID,Total", "--include-formulas", "--json",
        env={"AGW_HOME": agw_home},
    )
    data = json.loads(result.stdout)
    assert data["formula_mode"] == "value_and_formula"
    assert data["rows"][0][0] == {"value": "A-1", "formula": None}
    assert data["rows"][0][1] == {"value": None, "formula": "=1+1"}
    assert data["cached_values_may_be_stale"] is True


def test_cli_read_range_formulas_and_structural_validation(table_book, agw_home):
    ranged = run_agw(
        "office", "read-range", table_book, "--sheet", "Data",
        "--range", "C2:D2", "--formulas", "--json",
        env={"AGW_HOME": agw_home},
    )
    range_data = json.loads(ranged.stdout)
    assert range_data["rows"][0][1] == {"value": None, "formula": "=1+1"}

    checked = run_agw(
        "office", "validate-formulas", table_book, "--json",
        env={"AGW_HOME": agw_home},
    )
    formula_data = json.loads(checked.stdout)
    assert formula_data["validation"] == "structural_only"
    assert formula_data["calculation_performed"] is False
    assert formula_data["formula_count"] == 1
    assert formula_data["missing_cached_values"] == 1
    assert formula_data["formulas"][0]["cell"] == "D2"
    assert formula_data["formulas"][0]["formula"] == "=1+1"


def test_formula_validation_does_not_misclassify_structured_references(
        table_book, agw_home):
    wb = openpyxl.load_workbook(table_book)
    wb["Data"]["D2"] = "=[@Hours]*[@Rate]"
    wb.save(table_book)
    wb.close()
    data = office_excel.validate_formulas(table_book)
    assert data["formula_count"] == 1
    assert data["external_references"] == 0
    assert data["valid"] is True
    assert office_excel._coerce({
        "$formula": "=[@Hours]*[@Rate]"
    }) == "=[@Hours]*[@Rate]"


def test_cli_append_table_row_accepts_stdin_json_with_spaces(table_book, agw_home):
    # Windows PowerShell commonly prefixes native pipeline input with a UTF-8
    # BOM. Exercise that exact byte stream as a permanent regression case.
    payload = "\ufeff" + json.dumps({
        "ID": "A-2", "Status": "Needs review", "Due": "2026-08-01",
    })
    result = run_agw(
        "office", "append-table-row", table_book, "--table", "Orders",
        "--row-json", "-", "--coerce-iso-dates", "--json",
        env={"AGW_HOME": agw_home}, input_text=payload,
    )
    data = json.loads(result.stdout)
    assert data["appended"] == 1

    readback = run_agw(
        "office", "read-table", table_book, "--table", "Orders", "--json",
        env={"AGW_HOME": agw_home},
    )
    assert json.loads(readback.stdout)["rows"][1] == [
        "A-2", "Needs review", "2026-08-01", "=1+1",
    ]


def test_cli_ensure_table_stdin_dry_run_then_create(empty_book, agw_home):
    before = store.file_sha256(empty_book)
    args = (
        "office", "ensure-table", empty_book, "--sheet", "Records",
        "--table", "RecordsTable", "--headers-json", "-",
        "--create-sheet", "--expected-file-hash", before, "--json",
    )
    dry = run_agw(*args[:-1], "--dry-run", "--json", env={"AGW_HOME": agw_home},
                  input_text='["RecordID","Status"]')
    dry_data = json.loads(dry.stdout)
    assert dry_data["dry_run"] is True
    assert dry_data["before_hash"] == dry_data["after_hash"] == before
    assert store.list_versions(empty_book) == []

    created = run_agw(*args, env={"AGW_HOME": agw_home},
                      input_text='["RecordID","Status"]')
    data = json.loads(created.stdout)
    assert data["changed"] == 1
    assert data["sheet"] == "Records"
    assert data["range"] == "A1:B1"


def test_cli_uniqueness_conflict_is_structured(table_book, agw_home):
    result = run_agw(
        "office", "append-table-row", table_book, "--table", "Orders",
        "--row-json", '{"ID":"A-1","Status":"Closed"}',
        "--unique-column", "ID", "--json", env={"AGW_HOME": agw_home},
        check=False,
    )
    assert result.returncode == 3
    data = json.loads(result.stderr)
    assert data["ok"] is False
    assert data["error"]["code"] == "uniqueness_conflict"
    assert data["error"]["details"]["columns"] == ["ID"]


def test_end_to_end_tracker_replacement_workflow(tmp_path, agw_home):
    original = tmp_path / "tracker.xlsx"
    replacement = tmp_path / "tracker.replacement.xlsx"
    openpyxl.Workbook().save(original)
    original_hash = store.file_sha256(str(original))
    shutil.copy2(original, replacement)

    inspected = run_agw(
        "office", "info", str(original), "--json",
        env={"AGW_HOME": agw_home},
    )
    assert json.loads(inspected.stdout)["hash"] == original_hash

    projects = run_agw(
        "office", "ensure-table", str(replacement), "--sheet", "Projects",
        "--table", "Projects", "--headers-json", '["Project ID","Status"]',
        "--create-sheet", "--expected-file-hash",
        store.file_sha256(str(replacement)), "--json",
        env={"AGW_HOME": agw_home},
    )
    assert json.loads(projects.stdout)["table"] == "Projects"

    columns = json.dumps([
        {"name": "Date", "number_format": "yyyy-mm-dd"},
        {"name": "Hours", "number_format": "0.00"},
        {"name": "Rate", "number_format": "0.00"},
        {"name": "Total", "formula": "=B2*C2", "number_format": "0.00"},
    ])
    worklog = run_agw(
        "office", "ensure-table", str(replacement), "--sheet", "WorkLog",
        "--table", "WorkLog", "--headers-json",
        '["Date","Hours","Rate","Total"]', "--columns-json", columns,
        "--create-sheet", "--expected-file-hash",
        store.file_sha256(str(replacement)), "--json",
        env={"AGW_HOME": agw_home},
    )
    assert json.loads(worklog.stdout)["table"] == "WorkLog"

    appended = run_agw(
        "office", "append-table-row", str(replacement), "--table", "WorkLog",
        "--row-json", '{"Date":"2026-07-29","Hours":2,"Rate":150}',
        "--unique-column", "Date", "--coerce-iso-dates",
        "--expected-file-hash", store.file_sha256(str(replacement)), "--json",
        env={"AGW_HOME": agw_home},
    )
    assert json.loads(appended.stdout)["changed"] == 1

    formulas = run_agw(
        "office", "read-table", str(replacement), "--table", "WorkLog",
        "--include-formulas", "--json", env={"AGW_HOME": agw_home},
    )
    formula_data = json.loads(formulas.stdout)
    assert formula_data["rows"][0][3]["formula"] == "=B2*C2"
    replacement_hash = store.file_sha256(str(replacement))

    archived = run_agw(
        "archive", str(original), "--json", env={"AGW_HOME": agw_home},
    )
    archive_data = json.loads(archived.stdout)
    assert archive_data[0]["sha256"] == original_hash
    assert not original.exists()

    moved = run_agw(
        "move", str(replacement), str(original), "--json",
        env={"AGW_HOME": agw_home},
    )
    move_data = json.loads(moved.stdout)
    assert move_data["dest"] == str(original)
    assert store.file_sha256(str(original)) == replacement_hash
    assert office_tx.transaction_status() == []
    assert not list(tmp_path.glob(".agw-office-*"))


def test_word_outline_and_patch_if_dependency_present(tmp_path, agw_home):
    docx = pytest.importorskip("docx")
    path = tmp_path / "memo.docx"
    document = docx.Document()
    document.add_heading("Overview", level=1)
    document.add_paragraph("Original text.")
    document.save(path)

    view = office_word.outline(str(path))
    assert view["preservation"] == {"safe_to_mutate": True, "risks": []}
    paragraph_id = view["blocks"][1][0]
    result = office_word.patch(
        str(path),
        [{"op": "replace_block", "id": paragraph_id, "text": "Revised text."}],
        expected_sha256=view["hash"],
    )
    assert result["patched"] == 1
    assert docx.Document(path).paragraphs[1].text == "Revised text."

    updated = office_word.outline(str(path))
    updated_id = updated["blocks"][1][0]
    cli = run_agw(
        "office", "patch", str(path), "--ops-json", "-",
        "--expected-file-hash", updated["hash"], "--json",
        env={"AGW_HOME": agw_home},
        input_text=json.dumps([{
            "op": "replace_block", "id": updated_id,
            "text": "PowerShell-safe revision with spaces.",
        }]),
    )
    assert json.loads(cli.stdout)["patched"] == 1
    assert docx.Document(path).paragraphs[1].text == \
        "PowerShell-safe revision with spaces."
